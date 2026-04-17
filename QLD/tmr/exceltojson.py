import tkinter as tk
from tkinter import filedialog
import openpyxl
import json
import os
import datetime

def select_excel_file():
    """Opens a file dialog for the user to select an Excel file."""
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window
    file_path = filedialog.askopenfilename(
        title="Select an Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm")]
    )
    return file_path

def datetime_serializer(obj):
    """Helper function to serialize datetime objects for JSON."""
    if isinstance(obj, (datetime.datetime, datetime.date, datetime.time)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} is not serializable")

def convert_excel_to_json():
    # 1. Ask the user to select the file
    print("Please select an Excel file from the dialog...")
    file_path = select_excel_file()
    
    if not file_path:
        print("No file selected. Exiting.")
        return

    print(f"Processing file: {file_path}")
    
    # 2. Load the workbook
    # data_only=True gets the actual values, not the formulas
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    excel_data = {}

    # 3. Iterate through all sheets and extract values
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_rows = []
        
        # values_only=True instantly extracts the cell values instead of cell objects
        for row in sheet.iter_rows(values_only=True):
            sheet_rows.append(list(row))
        
        # Only add the sheet to our JSON if it contains data
        if sheet_rows:
            excel_data[sheet_name] = sheet_rows

    # 4. Construct the final JSON payload
    output_data = {
        "metadata": {
            "source_file": os.path.basename(file_path),
            "instruction": "This JSON file contains the exact data values extracted from the source Excel file organized by sheet name."
        },
        "data": excel_data
    }

    # 5. Save to a JSON file
    output_filename = os.path.splitext(os.path.basename(file_path))[0] + "_data.json"
    
    try:
        # Use default=datetime_serializer to handle any Excel date/time formats properly
        with open(output_filename, 'w', encoding='utf-8') as json_file:
            json.dump(output_data, json_file, indent=4, default=datetime_serializer)
        print(f"Success! Excel data saved to: {output_filename}")
    except Exception as e:
        print(f"Error saving JSON file: {e}")

if __name__ == "__main__":
    convert_excel_to_json()